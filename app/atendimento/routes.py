# app/atendimento/routes.py
from weasyprint import HTML

from flask import Blueprint, render_template, request, redirect, url_for, session, flash,make_response

from datetime import datetime, time
from werkzeug.utils import secure_filename
import os
from ..models import Obra, db, ComunicacaoObra, VistoriaImovel, FotoVistoria
from ..utils import registrar_acao
from ..services.bunny import upload_bunny
from flask_login import current_user
from sqlalchemy.orm import joinedload
from dateutil import parser  # se ainda não tiver
from flask import send_file  # se ainda não estiver
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader


from reportlab.lib.colors import black
from textwrap import wrap
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from dateutil import parser
import base64
from PIL import Image
from weasyprint import HTML
import io
from dotenv import load_dotenv
from app.fotos.utils import upload_foto_bunny  # vamos criar isso já já
from app.models import FotoVistoria
from app.fotos.upload_bunny import upload_foto_vistoria
from app.fotos.bunny import upload_foto_vistoria
from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from app.models import db, ComunicacaoObra, VistoriaImovel, Obra, FotoVistoria
from app.fotos.bunny import upload_foto_vistoria
from dateutil import parser
import base64
import requests
from io import BytesIO
from flask import current_app  # <-- adicione
import re

atendimento_bp = Blueprint("atendimento", __name__)
load_dotenv()




def parse_hora(hora_str):
    try:
        return datetime.strptime(hora_str, "%H:%M").time()
    except ValueError:
        return datetime.strptime(hora_str, "%H:%M:%S").time()

atendimento_bp = Blueprint('atendimento', __name__, template_folder='../../templates/atendimento')

@atendimento_bp.route("/")
def dashboard_unificado():
    if "user_id" not in session:
        return redirect(url_for("auth.login"))

    cargo = session.get("cargo")
    user_id = session.get("user_id")
    registros = []

    # 👇 OBRAS VISÍVEIS PARA O USUÁRIO (vistoriador = só a obra vinculada)
    if cargo == "vistoriador":
        from app.models import User  # necessário se current_user não está funcionando
        usuario = User.query.get(user_id)
        obras_ids = [usuario.obra_id] if usuario and usuario.obra_id else []
    else:
        obras_ids = [obra.id for obra in Obra.query.all()]

    # 🗂️ Buscar os registros conforme a obra_id permitida
    comunicacoes = ComunicacaoObra.query.filter(ComunicacaoObra.obra_id.in_(obras_ids)).all()
    vistorias = VistoriaImovel.query.filter(VistoriaImovel.obra_id.in_(obras_ids)).options(joinedload(VistoriaImovel.usuario)).all()

    # 🧽 Evita duplicação: só mostra comunicações sem vistoria associada
    ids_com_com_vistoria = [v.comunicacao_id for v in vistorias if v.comunicacao_id]

    for c in comunicacoes:
        if c.id not in ids_com_com_vistoria:
            registros.append({
                "id": c.id,
                "nome": c.nome,
                "rua": c.endereco,
                "obra": c.obra,
                "data_envio": c.data_envio,
                "finalizada": False,
                "comunicador": f"{c.usuario.nome} (Admin)" if c.usuario and c.usuario.cargo == "admin" else c.usuario.nome if c.usuario else "—"
            })

    for v in vistorias:
        registros.append({
    "id": v.id,
    "nome": v.nome_responsavel or (v.comunicacao.nome if v.comunicacao else "Sem nome"),
    "rua": v.rua or (v.comunicacao.endereco if v.comunicacao else "—"),
    "obra": v.obra,

    # 👉 1ª tentativa
    "data_envio": datetime.combine(v.data_1, v.hora_1) if v.data_1 and v.hora_1 else v.data_1,

    # 👉 2ª tentativa
    "segunda_tentativa": datetime.combine(v.data_2, v.hora_2) if v.data_2 and v.hora_2 else None,

    # 👉 3ª tentativa
    "terceira_tentativa": datetime.combine(v.data_3, v.hora_3) if v.data_3 and v.hora_3 else None,

    "finalizada": v.finalizada,
    "comunicador": (
        f"{v.comunicacao.usuario.nome} (Admin)" if v.comunicacao and v.comunicacao.usuario and v.comunicacao.usuario.cargo == "admin"
        else v.comunicacao.usuario.nome if v.comunicacao and v.comunicacao.usuario
        else "—"
    ),
    "vistoriador": (
        f"{v.usuario.nome} (Admin)" if v.usuario and v.usuario.cargo == "admin"
        else v.usuario.nome if v.usuario else "—"
    ),
})

    registros.sort(
        key=lambda r: datetime.combine(r["data_envio"], time.min) if r.get("data_envio") else datetime.min,
        reverse=True
    )

    return render_template("atendimento/dashboard.html",
                           registros=registros,
                           total_comunicacoes=len(comunicacoes),
                           total_vistorias=len(vistorias),
                           total_registros=len(registros))




@atendimento_bp.app_template_filter('getattr_safe')
def getattr_safe(obj, attr):
    return getattr(obj, attr, '')

@atendimento_bp.route("/<int:id>/delete", methods=["POST"])
def deletar_atendimento(id):
    if "user_id" not in session:
        return redirect(url_for("auth.login"))

    if session.get("cargo") != "admin":
        flash("Apenas administradores podem excluir atendimentos.", "danger")
        return redirect(url_for("atendimento.dashboard_unificado"))

    vistoria = VistoriaImovel.query.get_or_404(id)
    comunicacao = vistoria.comunicacao

    for foto in vistoria.fotos:
        db.session.delete(foto)

    db.session.delete(vistoria)
    if comunicacao:
        db.session.delete(comunicacao)

    db.session.commit()

    registrar_acao(
        tipo_acao="exclusão",
        entidade="VistoriaImovel",
        entidade_id=vistoria.id,
        usuario_id=session["user_id"],
        observacao=f"Atendimento vinculado à comunicação #{comunicacao.id if comunicacao else '—'} foi excluído."
    )

    if comunicacao:
        registrar_acao(
            tipo_acao="exclusão",
            entidade="ComunicacaoObra",
            entidade_id=comunicacao.id,
            usuario_id=session["user_id"],
            observacao=f"Comunicação associada à vistoria #{vistoria.id} foi excluída."
        )

    flash(f"❌ Atendimento #{id} excluído com sucesso!", "success")
    return redirect(url_for("atendimento.dashboard_unificado"))

@atendimento_bp.route("/nova")
def nova_comunicacao_vistoria():
    if "user_id" not in session:
        return redirect(url_for("auth.login"))
    
    obras = Obra.query.all()
    return render_template(
    "atendimento/formulario.html",
    obras=obras,
    comunicacao=None,
    vistoria=None
)


@atendimento_bp.route("/criar", methods=["POST"])
def criar_atendimento():
    if "user_id" not in session:
        flash("⚠️ Sessão inválida. Faça login novamente.", "danger")
        return redirect(url_for("auth.login"))

    nome = request.form.get("nome")
    endereco = request.form.get("endereco")
    obra_id = request.form.get("obra_id")

    if not nome or not endereco or not obra_id:
        flash("⚠️ Nome, endereço e obra são obrigatórios!", "danger")
        return redirect(url_for("atendimento.nova_comunicacao_vistoria"))

    # 1) Comunicação
    nova_comunicacao = ComunicacaoObra(
        nome=nome,
        cpf=request.form.get("cpf"),
        telefone=request.form.get("telefone"),
        endereco=endereco,
        numero=request.form.get("numero"),
        bairro=request.form.get("bairro"),
        comunicado=request.form.get("comunicado"),
        economia=request.form.get("economia"),
        tipo_imovel=request.form.get("tipo_imovel"),
        data_envio=datetime.now(),
        usuario_id=session["user_id"],
        obra_id=obra_id
    )
    db.session.add(nova_comunicacao)
    db.session.flush()  # para ter nova_comunicacao.id

    # 2) Vistoria inicial
    nova_vistoria = VistoriaImovel(
        comunicacao_id=nova_comunicacao.id,
        data_1=datetime.now().date(),
        hora_1=datetime.now().time(),
        obra_id=obra_id,
        usuario_id=session["user_id"] if session.get("cargo") == "vistoriador" else None
    )
    db.session.add(nova_vistoria)
    db.session.commit()  # garante nova_vistoria.id

    # 3) Logs
    registrar_acao(
        tipo_acao="criação",
        entidade="ComunicacaoObra",
        entidade_id=nova_comunicacao.id,
        usuario_id=session["user_id"],
        observacao=f"Comunicação criada para '{nova_comunicacao.nome}' na obra ID {obra_id}."
    )
    registrar_acao(
        tipo_acao="criação",
        entidade="VistoriaImovel",
        entidade_id=nova_vistoria.id,
        usuario_id=session["user_id"],
        observacao=f"Vistoria inicial registrada para comunicação ID {nova_comunicacao.id}."
    )

    # 4) Upload de fotos (se houver) para BunnyCDN
    fotos = request.files.getlist("fotos")
    if fotos:
        from app.fotos.bunny import upload_foto_vistoria  # garante import correto
        obra = Obra.query.get(obra_id)
        obra_nome = obra.nome if obra else "SemObra"

        enviados = 0
        for foto in fotos:
            if foto and foto.filename:
                try:
                    resultado = upload_foto_vistoria(
                        obra_nome=obra_nome,
                        vistoria_id=nova_vistoria.id,
                        foto_file=foto
                    )
                    nova_foto = FotoVistoria(
                        vistoria_id=nova_vistoria.id,
                        url=resultado["url"],
                        titulo=resultado["nome"],
                        descricao=""
                    )
                    db.session.add(nova_foto)
                    enviados += 1
                except Exception as e:
                    print(f"❌ Erro ao enviar {foto.filename}: {e}")
                    flash(f"Erro ao enviar {foto.filename}: {e}", "danger")
        db.session.commit()
        if enviados:
            flash(f"✅ {enviados} foto(s) enviada(s) com sucesso.", "success")

    flash("✅ Atendimento criado com sucesso!", "success")
    return redirect(url_for("atendimento.atendimento_unificado", id=nova_vistoria.id))


@atendimento_bp.route("/<int:id>", methods=["GET", "POST"])
def atendimento_unificado(id):
    if "user_id" not in session:
        return redirect(url_for("auth.login"))

    cargo = session.get("cargo")
    vistoria = VistoriaImovel.query.get_or_404(id)
    comunicacao = vistoria.comunicacao
    obras = Obra.query.all()

    if request.method == "POST":
        # ======================= DEBUG DOS DADOS =======================
        print("====== DADOS RECEBIDOS DA VISTORIA ======")
        print("Soleira:", request.form.get("soleira"))
        print("Calçada:", request.form.get("calcada"))
        print("Uso:", request.form.get("uso"))
        print("Vínculo:", request.form.get("tipo_vinculo"))
        print("Responsável info:", request.form.get("responsavel_info"))
        print("Observação Geral:", request.form.get("observacao_geral"))
        assinatura_debug = request.form.get("assinatura")
        print("Assinatura (base64) começa com:", assinatura_debug[:30] if assinatura_debug else None)
        print("Tentativas:")
        for i in range(1, 4):
            print(f"  data_{i}:", request.form.get(f"data_{i}"))
            print(f"  hora_{i}:", request.form.get(f"hora_{i}"))
        print("================================================")
        # ===============================================================

        # Atualiza comunicação (se permitido)
        if comunicacao and cargo in ["admin", "comunicador"]:
            comunicacao.nome = request.form.get("nome") or comunicacao.nome
            comunicacao.cpf = request.form.get("cpf") or comunicacao.cpf
            comunicacao.telefone = request.form.get("telefone") or comunicacao.telefone
            comunicacao.endereco = request.form.get("endereco") or comunicacao.endereco
            comunicacao.numero = request.form.get("numero") or comunicacao.numero
            comunicacao.bairro = request.form.get("bairro") or comunicacao.bairro
            comunicacao.comunicado = request.form.get("comunicado") or comunicacao.comunicado
            comunicacao.economia = request.form.get("economia") or comunicacao.economia
            comunicacao.tipo_imovel = request.form.get("tipo_imovel") or comunicacao.tipo_imovel
            comunicacao.obra_id = request.form.get("obra_id") or comunicacao.obra_id

        # Atualiza vistoria
        vistoria.finalizada = bool(request.form.get("finalizada"))
        vistoria.obra_id = request.form.get("obra_id") or None
        vistoria.tipo_imovel = request.form.get("tipo_imovel")
        vistoria.soleira = request.form.get("soleira")
        vistoria.calcada = request.form.get("calcada")
        vistoria.uso = request.form.get("uso")
        vistoria.tipo_vinculo = request.form.get("tipo_vinculo")
        vistoria.responsavel_info = request.form.get("responsavel_info")
        vistoria.observacao_geral = request.form.get("observacao_geral")

        # Dados herdados da comunicação
        vistoria.nome_responsavel = request.form.get("nome") or (comunicacao.nome if comunicacao else None)
        vistoria.cpf_responsavel = request.form.get("cpf") or (comunicacao.cpf if comunicacao else None)
        vistoria.rua = request.form.get("endereco") or (comunicacao.endereco if comunicacao else None)
        vistoria.numero = request.form.get("numero") or (comunicacao.numero if comunicacao else None)
        vistoria.bairro = request.form.get("bairro") or (comunicacao.bairro if comunicacao else None)

        # Assinatura
        assinatura_data = request.form.get("assinatura")
        if assinatura_data and assinatura_data.startswith("data:image/png;base64,"):
            vistoria.assinatura_base64 = assinatura_data

        # Tentativas
        for i in range(1, 4):
            data = request.form.get(f"data_{i}")
            hora = request.form.get(f"hora_{i}")
            if data:
                try:
                    setattr(vistoria, f"data_{i}", datetime.strptime(data, "%Y-%m-%d").date())
                except ValueError:
                    flash(f"⚠️ Data inválida no campo data_{i}.", "danger")
            if hora:
                try:
                    hora_convertida = parser.parse(hora).time()
                    setattr(vistoria, f"hora_{i}", hora_convertida)
                except (ValueError, TypeError):
                    flash(f"⚠️ Hora inválida no campo hora_{i}.", "danger")

        vistoriador_assumiu = False
        if not vistoria.usuario_id and cargo == "vistoriador":
            vistoria.usuario_id = session["user_id"]
            vistoriador_assumiu = True

        db.session.commit()
    fotos = request.files.getlist("fotos")
    for foto in fotos:
        if foto and foto.filename:
            try:
                resultado = upload_foto_vistoria(
                    obra_nome=vistoria.obra.nome,
                    vistoria_id=vistoria.id,
                    foto_file=foto
                )
                nova_foto = FotoVistoria(
                    vistoria_id=vistoria.id,
                    url=resultado["url"],
                    titulo=resultado["nome"],
                    descricao=""
                )
                db.session.add(nova_foto)
            except Exception as e:
                print(f"Erro ao enviar foto: {e}")


        db.session.commit()
        registrar_acao("edição", "VistoriaImovel", vistoria.id, session["user_id"],
                       f"Vistoria editada no atendimento #{vistoria.id}.")
        if comunicacao:
            registrar_acao("edição", "ComunicacaoObra", comunicacao.id, session["user_id"],
                           f"Comunicação editada no atendimento #{comunicacao.id}.")
        if assinatura_data:
            registrar_acao("assinatura", "VistoriaImovel", vistoria.id, session["user_id"],
                           "Assinatura do morador adicionada ou atualizada.")

        flash("✅ Atendimento atualizado com sucesso!", "success")
        if vistoriador_assumiu:
            flash("📌 Agora você é o responsável por esta vistoria.", "info")

        return redirect(url_for("atendimento.dashboard_unificado"))

    # GET → renderização do formulário
    tentativas = {}
    for i in range(1, 4):
        tentativas[f"data_{i}"] = getattr(vistoria, f"data_{i}", None)
        tentativas[f"hora_{i}"] = getattr(vistoria, f"hora_{i}", None)


    if request.files.getlist("fotos"):
        fotos = request.files.getlist("fotos")
        for foto in fotos:
            if foto.filename:
                url = upload_foto_vistoria(
                    token=os.getenv("BUNNY_STORAGE_API_KEY"),
                    obra_nome=comunicacao.obra.nome,
                    vistoria_id=vistoria.id,
                    foto=foto
                )
                if url:
                    nova_foto = FotoVistoria(
                        titulo=foto.filename,
                        url=url,
                        descricao="",
                        vistoria_id=vistoria.id
                    )
                    db.session.add(nova_foto)
        db.session.commit()

    return render_template("atendimento/formulario.html",
                           vistoria=vistoria,
                           comunicacao=comunicacao,
                           obras=obras,
                           cargo=cargo,
                           tentativas=tentativas)



@atendimento_bp.route("/assumir_vistoria/<int:id>", methods=["POST"])
def assumir_vistoria(id):
    if "user_id" not in session:
        return redirect(url_for("auth.login"))

    vistoria = VistoriaImovel.query.filter_by(comunicacao_id=id).first()
    if not vistoria:
        flash("Vistoria não encontrada.", "danger")
        return redirect(url_for("atendimento.dashboard_unificado"))

    if vistoria.usuario_id:
        flash("Essa vistoria já foi assumida.", "warning")
        return redirect(url_for("atendimento.dashboard_unificado"))

    vistoria.usuario_id = session["user_id"]
    db.session.commit()
    flash("✅ Vistoria assumida com sucesso!", "success")
    return redirect(url_for("atendimento.dashboard_unificado"))

def campo(valor, placeholder="________________"):
    if isinstance(valor, (str, int, float)):
        return str(valor) if valor not in ["", "None", "N/A"] else placeholder
    if valor is None:
        return placeholder
    return str(valor)



@atendimento_bp.route("/relatorio/excel/<int:id>")
def gerar_excel_vistoria(id):
    if "user_id" not in session or session.get("cargo") not in ["admin", "vistoriador"]:
        flash("Acesso restrito a administradores e vistoriadores.", "danger")
        return redirect(url_for("auth.login"))

    vistoria = VistoriaImovel.query.get_or_404(id)

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    buffer = BytesIO()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laudo Vistoria"

    dados = [
        ["Campo", "Valor"],
        ["Data da Vistoria", f"{vistoria.data_1 or ''} {vistoria.hora_1 or ''}"],
        ["Responsável", vistoria.nome_responsavel or ''],
        ["CPF", vistoria.cpf_responsavel or ''],
        ["Vínculo", vistoria.tipo_vinculo or ''],
        ["Endereço", f"{vistoria.rua}, {vistoria.numero} - {vistoria.bairro}, {vistoria.municipio}"],
        ["Tipo de Imóvel", vistoria.tipo_imovel or ''],
        ["Soleira", vistoria.soleira or ''],
        ["Calçada", vistoria.calcada or ''],
        ["Obra", vistoria.obra.nome if vistoria.obra else ''],
        ["Observações", vistoria.observacoes or '']
    ]

    for row in dados:
        ws.append(row)
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 35

    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"laudo_vistoria_{id}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@atendimento_bp.route("/relatorio/comunicacoes/excel")
def exportar_comunicacoes_excel():
    if "user_id" not in session or session.get("cargo") not in ["admin", "vistoriador"]:
        flash("Acesso restrito a administradores e vistoriadores.", "danger")
        return redirect(url_for("auth.dashboard"))

    buffer = BytesIO()
    wb = Workbook()

    # 1. Aba Geral com todas as comunicações
    ws_geral = wb.active
    ws_geral.title = "Relatório Geral"
    ws_geral.append(["ID", "Nome", "Telefone", "Endereço", "Número", "Bairro", "Obra", "Data Envio"])

    comunicacoes = ComunicacaoObra.query.all()
    for c in comunicacoes:
        ws_geral.append([
            c.id,
            c.nome,
            c.telefone,
            c.endereco,
            c.numero,
            c.bairro,
            c.obra.nome if c.obra else "—",
            c.data_envio.strftime("%d/%m/%Y %H:%M") if c.data_envio else ""
        ])

    # 2. Criar aba para cada obra
    obras = Obra.query.all()
    for obra in obras:
        nome_aba = obra.nome[:31]  # Excel aceita no máx 31 caracteres no título
        ws_obra = wb.create_sheet(title=nome_aba)
        ws_obra.append(["ID", "Nome", "Telefone", "Endereço", "Número", "Bairro", "Data Envio"])

        comunicacoes_obra = ComunicacaoObra.query.filter_by(obra_id=obra.id).all()
        for c in comunicacoes_obra:
            ws_obra.append([
                c.id,
                c.nome,
                c.telefone,
                c.endereco,
                c.numero,
                c.bairro,
                c.data_envio.strftime("%d/%m/%Y %H:%M") if c.data_envio else ""
            ])

    # Finaliza e retorna
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="relatorio_comunicacoes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# Helpers para datas/horas seguras
def parse_date_safe(date_str):
    try:
        return parser.parse(date_str).date() if date_str else None
    except Exception:
        return None

def parse_time_safe(time_str):
    try:
        return parser.parse(time_str).time() if time_str else None
    except Exception:
        return None

@atendimento_bp.route("/atendimento/<int:id>/editar", methods=["GET", "POST"])
def editar_atendimento(id):
    comunicacao = ComunicacaoObra.query.get_or_404(id)
    vistoria = VistoriaImovel.query.filter_by(comunicacao_id=id).first()
    obras = Obra.query.all()

    if request.method == "POST":
        # 📞 Comunicação
        comunicacao.nome = request.form.get("nome")
        comunicacao.cpf = request.form.get("cpf")
        comunicacao.telefone = request.form.get("telefone")
        comunicacao.endereco = request.form.get("endereco")
        comunicacao.bairro = request.form.get("bairro")
        comunicacao.numero = request.form.get("numero")
        comunicacao.comunicado = request.form.get("comunicado")
        comunicacao.economia = request.form.get("economia")
        comunicacao.tipo_imovel = request.form.get("tipo_imovel")
        comunicacao.assinatura = request.form.get("assinatura")
        comunicacao.obra_id = request.form.get("obra_id")

        # 🏘️ Vistoria
        if not vistoria:
            vistoria = VistoriaImovel(comunicacao_id=id)
            db.session.add(vistoria)

        vistoria.uso = request.form.get("uso")
        vistoria.nome_responsavel = request.form.get("nome_responsavel")
        vistoria.cpf_responsavel = request.form.get("cpf_responsavel")
        vistoria.tipo_vinculo = request.form.get("tipo_vinculo")
        vistoria.municipio = request.form.get("municipio")
        vistoria.bairro = request.form.get("bairro")
        vistoria.rua = request.form.get("rua")
        vistoria.numero = request.form.get("numero_vistoria")
        vistoria.complemento = request.form.get("complemento")
        vistoria.celular = request.form.get("celular")
        vistoria.tipo_imovel = request.form.get("tipo_imovel")
        vistoria.soleira = request.form.get("soleira")
        vistoria.calcada = request.form.get("calcada")
        vistoria.observacoes = request.form.get("observacoes")
        vistoria.assinatura_base64 = request.form.get("assinatura_base64")

        # mantém obra da vistoria alinhada com a comunicação
        vistoria.obra_id = comunicacao.obra_id

        # ⏱️ Tentativas (date/time seguros)
        vistoria.data_1 = parse_date_safe(request.form.get("data_1"))
        vistoria.hora_1 = parse_time_safe(request.form.get("hora_1"))
        vistoria.data_2 = parse_date_safe(request.form.get("data_2"))
        vistoria.hora_2 = parse_time_safe(request.form.get("hora_2"))
        vistoria.data_3 = parse_date_safe(request.form.get("data_3"))
        vistoria.hora_3 = parse_time_safe(request.form.get("hora_3"))

        # 1º commit: garantir IDs atualizados
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception("Erro ao salvar:", exc_info=e)
            flash("❌ Erro ao salvar alterações.", "danger")
            return redirect(url_for("atendimento.editar_atendimento", id=id))

        # 📸 Upload de fotos (somente no POST)
        enviados_count = 0
        if "fotos" in request.files and vistoria and vistoria.obra_id:
            fotos = request.files.getlist("fotos") or []
            obra = Obra.query.get(vistoria.obra_id)
            obra_nome = obra.nome if obra else "SemObra"

            for idx, foto in enumerate(fotos):
                if not foto or not foto.filename:
                    continue
                try:
                    # usa seu service centralizado
                    resultado = upload_foto_vistoria(obra_nome, vistoria.id, foto)
                    titulo_foto = (request.form.get(f"titulo_foto_{idx}") or foto.filename).strip()


                    nova_foto = FotoVistoria(
                        vistoria_id=vistoria.id,
                        url=resultado["url"],
                        titulo=titulo_foto,
                        descricao=""
                    )
                    db.session.add(nova_foto)
                    enviados_count += 1
                except Exception as e:
                    current_app.logger.exception(f"Erro ao enviar {foto.filename}: {e}")
                    flash(f"❌ Erro ao enviar {foto.filename}: {e}", "danger")

            # commit único para todas as fotos
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                current_app.logger.exception("Erro ao salvar fotos:", exc_info=e)
                flash("❌ Erro ao salvar fotos.", "danger")

        if enviados_count:
            flash(f"✅ {enviados_count} foto(s) enviada(s) com sucesso.", "success")

        flash("✅ Atendimento atualizado com sucesso!", "success")
        return redirect(url_for("atendimento.editar_atendimento", id=id))

    # GET (edição): carrega SOMENTE as fotos vinculadas a esta vistoria
    fotos_vistoria = FotoVistoria.query.filter_by(vistoria_id=vistoria.id).order_by(FotoVistoria.id.desc()).all() if vistoria else []

    return render_template(
        "atendimento/editar_atendimento.html",
        comunicacao=comunicacao,
        vistoria=vistoria,
        obras=obras,
        fotos_vistoria=fotos_vistoria
    )




@atendimento_bp.route("/remover_vistoriador/<int:id>", methods=["POST"])
def remover_vistoriador(id):
    if "user_id" not in session or session["cargo"] != "admin":
        abort(403)

    vistoria = VistoriaImovel.query.get_or_404(id)
    vistoria.usuario_id = None
    db.session.commit()
    flash("Vistoriador removido com sucesso.", "success")
    return redirect(url_for("atendimento.editar_atendimento", id=vistoria.comunicacao_id))



@atendimento_bp.route('/relatorio/pdf/<int:id>')
def gerar_laudo_weasy(id):
    vistoria = VistoriaImovel.query.get_or_404(id)
    comunicacao = vistoria.comunicacao
    obra = vistoria.obra
    usuario = vistoria.usuario

    # 🧼 Nome da obra com underline
    obra_nome_sanitizado = obra.nome.replace(" ", "_") if obra and obra.nome else "Obra"

    public_base = current_app.config.get("BUNNY_PUBLIC_BASE")

    fotos_db = FotoVistoria.query.filter_by(vistoria_id=vistoria.id).all()
    fotos_base64 = []

    for f in fotos_db:
        try:
            # 1) normaliza domínio para a sua Pull Zone oficial
            url = (f.url or "").replace(" ", "%20")
            if public_base:
                url = re.sub(r"^https?://[^/]+", public_base, url)

            # 2) tenta baixar com cabeçalhos amigáveis
            headers = {
                "User-Agent": "Mozilla/5.0",
                "Referer": public_base or "http://localhost"
            }
            resp = requests.get(url, headers=headers, timeout=30)
            resp.raise_for_status()

            img_b64 = base64.b64encode(resp.content).decode("utf-8")
            fotos_base64.append({
                "titulo": f.titulo or (url.split("/")[-1] or "Foto"),
                "url": url,
                "base64": img_b64
            })
        except Exception as e:
            print(f"Erro ao baixar imagem {f.url}: {e}")

    html = render_template(
        "laudo_template/laudo_vistoria.html",
        vistoria=vistoria,
        comunicacao=comunicacao,
        obra=obra,
        usuario=usuario,
        fotos=fotos_base64,
        nome_morador=comunicacao.nome if comunicacao else "—",
        cpf_morador=comunicacao.cpf if comunicacao else "—",
        rg_vistoriador=usuario.rg if usuario else "—"
    )

    pdf_io = BytesIO()
    HTML(string=html).write_pdf(pdf_io)
    pdf_io.seek(0)

    return send_file(pdf_io, as_attachment=True, download_name=f"laudo_vistoria_{id}.pdf", mimetype="application/pdf")